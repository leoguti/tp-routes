#!/usr/bin/env python3
"""
Convierte el Excel de resoluciones de los pasantes (data/rutas_resoluciones.xlsx)
al JSON que consume scripts/seed.js (formato descrito en FORMATO_JSON.md).

Uso:
    python3 scripts/convert_xlsx_to_json.py

Genera: data/rutas_boyaca.json

Reglas aplicadas:
  - Hoja "Lista de empresas" -> sección `operadores` (solo nombre por ahora).
  - Hoja "Rutas":
      - La cabecera está en la fila 3. Las filas con ID/Origen/Destino inician
        un grupo; las filas siguientes sin ID heredan origen/destino/vía.
      - Empresas explícitas en cada fila; resoluciones por fila.
      - "Directo" como vía -> via: [].
      - "Vía" literal como vía -> fila basura, se ignora.
      - Se agrupan múltiples resoluciones de la misma (origen,destino,via,empresa)
        en un solo objeto de ruta con `resoluciones` = [..].
      - Textos de resolución se parsean intentando extraer número y año:
        patrones como "RESOLUCIÓN 493 DE 1993", "RES 493 15-12-93",
        "RES 2634 DEL 31 DE MAYO DE 1993", etc. Si no se puede, queda
        texto_original con number null para revisión manual.
"""

import json
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
XLSX = REPO / "data" / "rutas_resoluciones.xlsx"
OUT  = REPO / "data" / "rutas_boyaca.json"
REGION = "boyaca"

MESES = {
    "enero":1,"febrero":2,"marzo":3,"abril":4,"mayo":5,"junio":6,
    "julio":7,"agosto":8,"septiembre":9,"octubre":10,"noviembre":11,"diciembre":12,
}


def norm(t):
    """Normaliza texto: trim + lowercase + sin tildes + espacios colapsados."""
    if t is None:
        return ""
    t = str(t).strip()
    t = unicodedata.normalize("NFD", t)
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    t = re.sub(r"\s+", " ", t).lower()
    return t


def parse_via(raw):
    """'Nobsa' -> ['Nobsa']; 'Tunja - Duitama' -> ['Tunja','Duitama'];
    'Directo' -> []; None/'' -> [] (ruta directa)."""
    if not raw:
        return []
    s = str(raw).strip()
    if norm(s) == "directo":
        return []
    parts = [p.strip() for p in s.split("-")]
    return [p for p in parts if p]


def parse_resolucion(raw):
    """Intenta extraer (numero, fecha_iso). Retorna dict con campos conocidos
    y siempre mantiene `texto_original` para trazabilidad."""
    if not raw:
        return None
    texto = str(raw).strip()
    # normalizar acento grave (Ò) a tilde/nada para comparar
    t_norm = norm(texto)

    # Caso "RESOLUCION NO. 2634 DEL 31 DE MAYO DE 1993"
    m = re.search(
        r"(\d{3,5})\s+del?\s+(\d{1,2})\s+de\s+([a-z]+)\s+de\s+(\d{4})",
        t_norm,
    )
    if m:
        num, dia, mes_txt, anio = m.groups()
        mes = MESES.get(mes_txt)
        fecha = f"{anio}-{mes:02d}-{int(dia):02d}" if mes else None
        return {
            "numero": num,
            "fecha": fecha,
            "texto_original": texto,
            "pdf_url": None,
        }

    # Caso "RES 493 15-12-93" -> número 493, fecha dd-mm-yy
    m = re.search(r"(\d{3,5}).{0,8}?(\d{1,2})-(\d{1,2})-(\d{2,4})", t_norm)
    if m:
        num, dia, mes, anio = m.groups()
        anio = int(anio)
        if anio < 100:
            anio = 1900 + anio if anio > 50 else 2000 + anio
        fecha = f"{anio}-{int(mes):02d}-{int(dia):02d}"
        return {
            "numero": num,
            "fecha": fecha,
            "texto_original": texto,
            "pdf_url": None,
        }

    # Caso "RESOLUCIÓN 493 DE 1993" o "RESOLUCIÒN 0104 DE 1999" (acento grave)
    m = re.search(r"(\d{3,5})\s+de\s+(\d{4})", t_norm)
    if m:
        num, anio = m.groups()
        return {
            "numero": num,
            "fecha": f"{anio}-01-01",
            "texto_original": texto,
            "pdf_url": None,
        }

    # Caso solo "RESOLUCIÓN 1234"
    m = re.search(r"(\d{3,5})", t_norm)
    if m:
        return {
            "numero": m.group(1),
            "fecha": None,
            "texto_original": texto,
            "pdf_url": None,
        }

    # No pude extraer número — dejo texto_original sin numero estructurado,
    # y uso el texto completo como numero (algo es mejor que nada).
    return {
        "numero": texto[:80],
        "fecha": None,
        "texto_original": texto,
        "pdf_url": None,
    }


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    # --- Operadores -----
    ws_ops = wb["Lista de empresas"]
    operadores_raw = []
    for (val,) in ws_ops.iter_rows(values_only=True):
        if val and str(val).strip().upper() != "EMPRESAS":
            operadores_raw.append(str(val).strip())

    # --- Rutas -----
    ws = wb["Rutas"]
    rows = list(ws.iter_rows(values_only=True))
    data_rows = rows[3:]  # saltar 2 filas vacías + header

    current = {"origen": None, "destino": None, "via_raw": None}
    parsed_rows = []
    last_empresa = None
    basura_skipped = 0

    for r in data_rows:
        if len(r) < 8:
            continue
        _, _, rid, origen, destino, empresa, via, resolucion = r[:8]

        if rid:  # fila cabecera de grupo
            current = {
                "origen": (str(origen).strip() if origen else None),
                "destino": (str(destino).strip() if destino else None),
                "via_raw": (str(via).strip() if via else None),
            }

        # fila basura donde todo es literal "Origen"/"Destino"/"Empresa"/"Vía"
        if current["origen"] and norm(current["origen"]) == "origen":
            basura_skipped += 1
            continue

        # vía puede venir en la propia fila si no estaba en cabecera
        via_raw = current["via_raw"] or (str(via).strip() if via else None)

        if empresa:
            last_empresa = str(empresa).strip()
            empresa_efectiva = last_empresa
        elif resolucion and last_empresa:
            # misma empresa anterior, otra resolución (caso Autoboy)
            empresa_efectiva = last_empresa
        else:
            continue

        if not (current["origen"] and current["destino"] and empresa_efectiva):
            continue

        parsed_rows.append({
            "origen":     current["origen"],
            "destino":    current["destino"],
            "via":        parse_via(via_raw),
            "operador":   empresa_efectiva,
            "resolucion_texto": (str(resolucion).strip() if resolucion else None),
        })

    # --- Agrupar: una entrada por (origen, destino, via, operador) ---
    # acumulando resoluciones.
    groups = {}
    for r in parsed_rows:
        key = (
            norm(r["origen"]),
            norm(r["destino"]),
            tuple(norm(v) for v in r["via"]),
            norm(r["operador"]),
        )
        if key not in groups:
            groups[key] = {
                "origen": r["origen"],
                "destino": r["destino"],
                "via": r["via"],
                "operador": r["operador"],
                "resoluciones": [],
            }
        res = parse_resolucion(r["resolucion_texto"])
        if res:
            # evitar duplicar si aparece idéntica
            if not any(
                x.get("texto_original") == res.get("texto_original")
                for x in groups[key]["resoluciones"]
            ):
                groups[key]["resoluciones"].append(res)

    rutas = list(groups.values())

    # --- Operadores referenciados que no están en la hoja (por si acaso) ---
    ops_norm = {norm(o): o for o in operadores_raw}
    for r in rutas:
        if norm(r["operador"]) not in ops_norm:
            ops_norm[norm(r["operador"])] = r["operador"]
    operadores = [{"nombre": v} for _, v in sorted(ops_norm.items())]

    out = {
        "region": REGION,
        "operadores": operadores,
        "rutas": rutas,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"OK -> {OUT}")
    print(f"  operadores: {len(operadores)}")
    print(f"  rutas:      {len(rutas)}")
    print(f"  rutas con resoluciones: {sum(1 for r in rutas if r['resoluciones'])}")
    print(f"  rutas sin resolucion:   {sum(1 for r in rutas if not r['resoluciones'])}")
    print(f"  total resoluciones:     {sum(len(r['resoluciones']) for r in rutas)}")
    print(f"  filas basura ignoradas: {basura_skipped}")


if __name__ == "__main__":
    main()
